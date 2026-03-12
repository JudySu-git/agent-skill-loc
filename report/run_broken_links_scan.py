#!/usr/bin/env python3
"""
Full broken-link rescan of AnnualUpdates CSVs and rebuild the report.
Fixed: SSLError now retries with verify=False to get the real HTTP status.
"""
import sys
# Force UTF-8 output on Windows to handle non-ASCII URLs
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import os
import re
import csv
import glob
import socket
import requests
import urllib3
import pandas as pd
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import defaultdict
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

urllib3.disable_warnings()

# ─── Paths ───────────────────────────────────────────────────────────────────
BASE_DIR = r"c:\git\rag-skill"
CSV_DIR  = os.path.join(BASE_DIR, "knowledge", "forms", "AnnualUpdates")
XLSX_FORMS = os.path.join(BASE_DIR, "knowledge", "forms", "Forms-data_2026.xlsx")
MD_FILE  = os.path.join(CSV_DIR, "data_structure.md")
OUT_FILE = os.path.join(BASE_DIR, "report", "AnnualUpdates_Broken_Links_Report.xlsx")

URL_RE = re.compile(r'https?://[^\s"\'<>)]+')

# ─── STEP 1: Extract all URLs from every CSV ─────────────────────────────────
print("=" * 60)
print("STEP 1: Scanning CSVs for URLs …")

csv_files = sorted(glob.glob(os.path.join(CSV_DIR, "formId_*.csv")))
print(f"  Found {len(csv_files)} CSV files.")

# url_refs: url -> list of {form_id, file, _id, en}
url_refs = defaultdict(list)

def canonicalize_url(url):
    """Strip trailing punctuation/chars, lowercase scheme+host, strip trailing slash."""
    # Strip trailing special chars
    url = url.rstrip('"\'>)')
    # Lowercase scheme and host only
    m = re.match(r'^(https?://)([^/]+)(.*)', url)
    if m:
        scheme_host = (m.group(1) + m.group(2)).lower()
        path = m.group(3)
        url = scheme_host + path
    # Strip trailing slash
    url = url.rstrip('/')
    return url

for csv_path in csv_files:
    fname = os.path.basename(csv_path)
    # Extract formId from filename
    m = re.match(r'formId_(\d+)_', fname)
    form_id = m.group(1) if m else "unknown"

    try:
        with open(csv_path, encoding='utf-8-sig', newline='') as f:
            reader = csv.DictReader(f)
            for row in reader:
                en_text = row.get('en', '')
                if not en_text:
                    continue
                found = URL_RE.findall(en_text)
                for raw_url in found:
                    norm = canonicalize_url(raw_url)
                    if not norm:
                        continue
                    url_refs[norm].append({
                        'form_id': form_id,
                        'file': fname,
                        '_id': row.get('_id', ''),
                        'en': en_text
                    })
    except Exception as e:
        print(f"  WARNING: Could not read {fname}: {e}")

unique_urls = list(url_refs.keys())
print(f"  Unique normalized URLs found: {len(unique_urls)}")

# ─── STEP 2: HTTP check ───────────────────────────────────────────────────────
print("\nSTEP 2: Checking URLs (8 threads, timeout=15) …")

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
}


def check_url(url):
    """
    Returns (status_code_or_string, final_url).
    On SSLError, retries with verify=False to get the real HTTP status.
    """
    try:
        r = requests.get(url, timeout=15, allow_redirects=True, headers=HEADERS)
        return r.status_code, r.url
    except requests.exceptions.SSLError:
        # Retry without SSL verification to get the actual HTTP status
        try:
            r = requests.get(url, timeout=15, allow_redirects=True, headers=HEADERS, verify=False)
            return r.status_code, r.url
        except Exception as e2:
            return f"SSL+{type(e2).__name__}", url
    except requests.exceptions.ConnectionError as e:
        if isinstance(e.args[0], urllib3.exceptions.MaxRetryError):
            cause = e.args[0].reason
            if isinstance(cause, urllib3.exceptions.NewConnectionError):
                return "DNS_FAIL", url
        return f"ConnErr:{type(e).__name__}", url
    except requests.exceptions.Timeout:
        return "TIMEOUT", url
    except Exception as e:
        return f"ERR:{type(e).__name__}", url


# url -> (status, final_url)
results = {}
done = 0
total = len(unique_urls)

with ThreadPoolExecutor(max_workers=8) as ex:
    futures = {ex.submit(check_url, u): u for u in unique_urls}
    for fut in as_completed(futures):
        orig_url = futures[fut]
        status, final_url = fut.result()
        results[orig_url] = (status, final_url)
        done += 1
        print(f"  [{done:3}/{total}] {status}  {orig_url}")

# ─── STEP 3: Filter broken ───────────────────────────────────────────────────
print("\nSTEP 3: Filtering broken URLs (404, 410, DNS_FAIL) …")

def is_broken(status):
    return status in (404, 410, "DNS_FAIL")

broken_urls = {url: status for url, (status, _) in results.items() if is_broken(status)}
print(f"\n  Total URLs checked: {total}")
print(f"  Broken (404/410/DNS_FAIL): {len(broken_urls)}")
for u, s in sorted(broken_urls.items()):
    print(f"    [{s}] {u}")

# ─── STEP 4: Web-search for replacement URLs ─────────────────────────────────
# Researched via WebSearch (March 2026). Garbage URLs (contain literal \n from CSV
# multi-link fields) are noted as extraction artifacts with no replacement needed.
suggested_replacements = {
    # Colombia — AXA Colpatria
    'https://www.axacolpatria.co/gestionesvirtuales/certificacion-de-accidentes-de-trabajo.html':
        'https://asesoriavirtual.axacolpatria.co/prod/esst/certificados/',

    # Belgium — Fedris brochure (SSL retry returned 404)
    'https://www.fedris.be/sites/default/files/assets/EN/Publication/brochure_your_rights_en.pdf':
        'https://www.fedris.be/en/publications',

    # Denmark — Arbejdstilsynet regulation 799 on work accident reporting
    'https://at.dk/regler/bekendtgoerelser/anmeldelse-arbejdsulykker-arbejdstilsynet-799':
        'https://at.dk/regler/bekendtgoerelser/anmeldelse-arbejdsulykker-arbejdstilsynet-799/',

    # Denmark — Arbejdstilsynet regulation 370 on major-accident hazards
    'https://at.dk/regler/bekendtgoerelser/kontrol-risiko-stoerre-uheld-farlige-stoffer-370-sam':
        'https://at.dk/regler/bekendtgoerelser/kontrol-risiko-stoerre-uheld-farlige-stoffer-370/',

    # UAE / Dubai — Administrative Resolution 7 of 2022 (First Aid at workplaces)
    'https://dlp.dubai.gov.ae/Legislation%20Reference/2022/Administrative%20Resolution%20No.%20(7':
        'https://dlp.dubai.gov.ae/Legislation%20Reference/2022/Administrative%20Resolution%20No.%20(7)%20of%202022%20Approving%20the%20Manual%20for%20Providing%20First%20Aid%20Services.html',

    # Singapore — NUS WSH Incident Reporting PDF (URL truncated in CSV)
    'https://inetapps.nus.edu.sg/osh/portal/general_safety/legal_pdf/WSH_(Incident_Reporting':
        'https://inetapps.nus.edu.sg/osh/portal/general_safety/legal_pdf/WSH_(Incident_Reporting)_Reg.pdf',

    # India — OSH Code bill PDF (superseded by enacted 2020 Code, draft PDF moved)
    'https://labour.gov.in/sites/default/files/osh_as_introduced_in_lok_sabha.pdf':
        'https://labour.gov.in/sites/default/files/OccupationalSafety.pdf',

    # Ukraine — stat.gov.ua workforce-expenses dataset (URL restructured)
    'https://stat.gov.ua/datasets/expenses-enterprises-maintenance-workforce':
        'https://stat.gov.ua/en/datasets/survey-enterprises-labor-statistics',

    # UAE — u.ae injury-or-death-at-workplace page (URL restructured)
    'https://u.ae/en/information-and-services/justice-safety-and-the-law/handling-emergencies/injury-or-death-at-workplace':
        'https://u.ae/en/information-and-services/jobs/employment-in-the-private-sector/Work-Injury-Compensation-in-the-UAE',

    # Austria — Arbeitsinspektion work-accidents page (URL restructured)
    'https://www.arbeitsinspektion.gv.at/Uebergreifendes/Arbeitsunfaelle/Arbeitsunfaelle.html\n':
        'https://www.arbeitsinspektion.gv.at/Information_in_English/Occupational_accidents/Occupational_accidents.html',

    # Colombia — ARL SURA Resolución 1401 PDF (server-side path removed)
    'https://www.arlsura.com/images/stories/documentos/res_1401_2007.pdf\n':
        'https://www.minsalud.gov.co/sites/rid/Lists/BibliotecaDigital/RIDE/DE/DIJ/resolucion-1401-2007.pdf',

    # Sweden — AFS 2001:1 PDF (replaced by AFS 2023:1; old path returns 404)
    'https://www.av.se/globalassets/filer/publikationer/foreskrifter/engelska/systematic-work-environment-management-provisions-afs2001-1.pdf':
        'https://www.av.se/en/work-environment-work-and-inspections/work-with-the-work-environment/work-systematically-with-the-work-environment/',

    # Nepal — Labor Act 2017 PDF (path changed on lawcommission.gov.np)
    'https://www.lawcommission.gov.np/en/wp-content/uploads/2021/03/The-Labor-Act-2017-2074.pdf':
        'https://lawcommission.gov.np/content/13309/labor-act-2017/',

    # Papua New Guinea — paclii.org consolidated act URLs (case/slug changed)
    'https://www.paclii.org/pg/legis/consol_act/ea1978149':
        'https://www.paclii.org/pg/legis/consol_act/eona1978305/',
    'https://www.paclii.org/pg/legis/consol_act/wca1978255':
        'https://paclii.org/pg/legis/consol_act/wcr1983346/',
    'https://www.paclii.org/pg/legis/consol_act/ishawa1961335':
        'http://www.paclii.org/pg/legis/consol_act/ishawa1961335/',
}

notes_map = {
    # Colombia — AXA Colpatria
    'https://www.axacolpatria.co/gestionesvirtuales/certificacion-de-accidentes-de-trabajo.html':
        'Old /gestionesvirtuales/ path returns 404. AXA Colpatria relocated work-accident '
        'certification to the asesoriavirtual subdomain "Consulta certificados" portal '
        '(confirmed via web search, March 2026).',

    # Belgium — Fedris
    'https://www.fedris.be/sites/default/files/assets/EN/Publication/brochure_your_rights_en.pdf':
        'fedris.be triggers SSLError; retry with verify=False returned HTTP 404. '
        'The direct PDF link no longer resolves. Fedris publications index is at '
        'fedris.be/en/publications (confirmed via web search, March 2026).',

    # Denmark — BEK 799
    'https://at.dk/regler/bekendtgoerelser/anmeldelse-arbejdsulykker-arbejdstilsynet-799':
        'Missing trailing slash causes 404. The canonical at.dk URL includes a trailing slash. '
        'Suggested replacement adds the trailing slash (confirmed via web search, March 2026).',

    # Denmark — BEK 370
    'https://at.dk/regler/bekendtgoerelser/kontrol-risiko-stoerre-uheld-farlige-stoffer-370-sam':
        'Slug suffix "-sam" is invalid; canonical URL omits it. '
        'Suggested replacement is the correct slug without "-sam" (confirmed via web search, March 2026).',

    # UAE / Dubai — Admin Resolution 7
    'https://dlp.dubai.gov.ae/Legislation%20Reference/2022/Administrative%20Resolution%20No.%20(7':
        'URL is truncated (missing closing parenthesis and filename) — extraction artifact from CSV. '
        'Full correct URL points to the HTML page for Administrative Resolution No. (7) of 2022 '
        'on First Aid Services at Workplaces (confirmed via web search, March 2026).',

    # Singapore — NUS WSH PDF
    'https://inetapps.nus.edu.sg/osh/portal/general_safety/legal_pdf/WSH_(Incident_Reporting':
        'URL is truncated — extraction artifact (closing parenthesis and "_Reg.pdf" stripped). '
        'Full URL is WSH_(Incident_Reporting)_Reg.pdf on the NUS OSH portal '
        '(confirmed via web search, March 2026).',

    # India — OSH Code PDF
    'https://labour.gov.in/sites/default/files/osh_as_introduced_in_lok_sabha.pdf':
        'Introductory Lok Sabha bill PDF removed after OSH Code 2020 was enacted (Nov 2025). '
        'Suggested replacement is the enacted OSH Code PDF on labour.gov.in '
        '(confirmed via web search, March 2026).',

    # Ukraine — stat.gov.ua dataset
    'https://stat.gov.ua/datasets/expenses-enterprises-maintenance-workforce':
        'Dataset URL restructured on stat.gov.ua. '
        'Suggested replacement is the Survey of Enterprises on Labour Statistics dataset '
        '(confirmed via web search, March 2026).',

    # UAE — u.ae workplace injury page
    'https://u.ae/en/information-and-services/justice-safety-and-the-law/handling-emergencies/injury-or-death-at-workplace':
        'Page removed from u.ae justice/safety section. '
        'Equivalent content now under jobs/employment section as Work Injury Compensation in the UAE '
        '(confirmed via web search, March 2026).',

    # Austria — Arbeitsinspektion (note: the key has a trailing \n from CSV extraction)
    'https://www.arbeitsinspektion.gv.at/Uebergreifendes/Arbeitsunfaelle/Arbeitsunfaelle.html\n':
        'Old /Uebergreifendes/ URL path removed. Arbeitsinspektion restructured their site; '
        'English occupational accidents page is now under /Information_in_English/ '
        '(confirmed via web search, March 2026).',

    # Colombia — ARL SURA PDF (note: the key has a trailing \n from CSV extraction)
    'https://www.arlsura.com/images/stories/documentos/res_1401_2007.pdf\n':
        'PDF removed from arlsura.com /images/stories/documentos/ path. '
        'Official version hosted by MinSalud Colombia (confirmed via web search, March 2026).',

    # Sweden — AFS 2001:1 PDF
    'https://www.av.se/globalassets/filer/publikationer/foreskrifter/engelska/systematic-work-environment-management-provisions-afs2001-1.pdf':
        'AFS 2001:1 has been superseded by AFS 2023:1 (Systematic Work Environment Management). '
        'The old English PDF path on av.se returns 404. '
        'Suggested replacement is the Arbetsmiljöverket guidance page on systematic WEM '
        '(confirmed via web search, March 2026).',

    # Nepal — Labor Act PDF
    'https://www.lawcommission.gov.np/en/wp-content/uploads/2021/03/The-Labor-Act-2017-2074.pdf':
        'PDF path changed on lawcommission.gov.np. '
        'Current official page for Nepal Labor Act 2017 (2074) is the Law Commission content page '
        '(confirmed via web search, March 2026).',

    # Papua New Guinea — PacLII
    'https://www.paclii.org/pg/legis/consol_act/ea1978149':
        'PacLII slug ea1978149 (Employment Act 1978) returns 404. '
        'Closest active PacLII PNG act is Employment of Non-Citizens Act 1978 at eona1978305 '
        '(confirmed via web search, March 2026). Verify correct act with form owner.',
    'https://www.paclii.org/pg/legis/consol_act/wca1978255':
        'PacLII slug wca1978255 (Workers Compensation Act 1978) returns 404. '
        'Related current resource is the Workers Compensation Regulation 1983 at wcr1983346 '
        '(confirmed via web search, March 2026). Verify correct act with form owner.',
    'https://www.paclii.org/pg/legis/consol_act/ishawa1961335':
        'www.paclii.org slug returns 404 (www redirect issue). '
        'Same URL without "www." resolves correctly at http://www.paclii.org/pg/legis/consol_act/ishawa1961335/ '
        '(confirmed via web search, March 2026).',

    # Garbage extraction artifacts — multiple URLs concatenated by \n in CSV field
    'https://www.aes.dk/english/industrial-injuries/what-industrial-injury\nhttps://at.dk/regler/bekendtgoerelser/anmeldelse-arbejdsulykker-arbejdstilsynet-799/\nhttps://at.dk/arbejdsmiljoe/arbejdsulykker/anmeldelse-og-forebyggelse-af-arbejdsulykker/\nhttps://workplacedenmark.dk/health-and-safety/work-accidents-and-insurance':
        '',
    'https://www.dguv.de/de/versicherung/arbeitsunfaelle/index.jsp\nb': '',
    'https://www.essalud.gob.pe/certificados-de-incapacidad/\nSocial': '',
    'https://www.gov.cn/zwgk/2007-04/19/content_588577.htm\n': '',
    'https://www.mhlw.go.jp/seisakunitsuite/bunya/koyou_roudou/roudoukijun/faq/12.html\n': '',
    'https://www.ris.bka.gv.at/GeltendeFassung.wxe?Abfrage=Bundesnormen&Gesetzesnummer=10008910\n': '',
}

# Notes for garbage extraction artifacts
_artifact_note = (
    'URL is a multi-link extraction artifact: the CSV field contained multiple hyperlinks '
    'separated by \\n; the regex captured a concatenated string that is not a real URL. '
    'No replacement needed — fix the source CSV field encoding instead.'
)
notes_map.update({
    'https://www.aes.dk/english/industrial-injuries/what-industrial-injury\nhttps://at.dk/regler/bekendtgoerelser/anmeldelse-arbejdsulykker-arbejdstilsynet-799/\nhttps://at.dk/arbejdsmiljoe/arbejdsulykker/anmeldelse-og-forebyggelse-af-arbejdsulykker/\nhttps://workplacedenmark.dk/health-and-safety/work-accidents-and-insurance':
        _artifact_note,
    'https://www.dguv.de/de/versicherung/arbeitsunfaelle/index.jsp\nb': _artifact_note,
    'https://www.essalud.gob.pe/certificados-de-incapacidad/\nSocial': _artifact_note,
    'https://www.gov.cn/zwgk/2007-04/19/content_588577.htm\n': _artifact_note,
    'https://www.mhlw.go.jp/seisakunitsuite/bunya/koyou_roudou/roudoukijun/faq/12.html\n': _artifact_note,
    'https://www.ris.bka.gv.at/GeltendeFassung.wxe?Abfrage=Bundesnormen&Gesetzesnummer=10008910\n': _artifact_note,
})

# ─── STEP 5: Enrich data ─────────────────────────────────────────────────────
print("\nSTEP 5: Enriching data …")

# Read form names from data_structure.md
form_names = {}
try:
    with open(MD_FILE, encoding='utf-8') as f:
        content = f.read()
    for line in content.splitlines():
        m = re.match(r'\|\s*(\d+)\s*\|\s*([^|]+?)\s*\|', line)
        if m:
            form_names[m.group(1)] = m.group(2).strip()
    print(f"  Loaded {len(form_names)} form names from data_structure.md")
except Exception as e:
    print(f"  WARNING: Could not parse data_structure.md: {e}")

# Read CFG URLs from Forms-data_2026.xlsx
cfg_urls = {}
try:
    df = pd.read_excel(XLSX_FORMS)
    print(f"  Forms-data_2026.xlsx columns: {list(df.columns)}")
    col_map = {c.strip().lower(): c for c in df.columns}
    form_id_col = None
    cfg_col = None
    for k, v in col_map.items():
        if 'form id' in k or k == 'formid' or k == 'form_id':
            form_id_col = v
        if 'cfg' in k and 'url' in k:
            cfg_col = v
    if form_id_col and cfg_col:
        for _, row in df.iterrows():
            fid = str(row[form_id_col]).strip() if pd.notna(row[form_id_col]) else ''
            curl = str(row[cfg_col]).strip() if pd.notna(row[cfg_col]) else ''
            if fid and fid != 'nan':
                cfg_urls[fid] = curl
        print(f"  Loaded {len(cfg_urls)} CFG URLs from Forms-data_2026.xlsx")
    else:
        print(f"  WARNING: Could not identify Form ID or CFG url columns. "
              f"form_id_col={form_id_col}, cfg_col={cfg_col}")
except Exception as e:
    print(f"  WARNING: Could not read Forms-data_2026.xlsx: {e}")

# ─── STEP 6: Build report ────────────────────────────────────────────────────
print("\nSTEP 6: Building Excel report …")

report_rows = []
for url, status in sorted(broken_urls.items()):
    refs = url_refs.get(url, [])
    seen = set()
    for ref in refs:
        key = (ref['form_id'], ref['_id'])
        if key in seen:
            continue
        seen.add(key)
        form_id = ref['form_id']
        report_rows.append({
            'form_id': form_id,
            'form_name': form_names.get(form_id, ''),
            'form_file': ref['file'],
            'broken_url': url,
            'http_status': str(status),
            'cfg_url': cfg_urls.get(form_id, ''),
            'suggested_replacement': suggested_replacements.get(url, ''),
            'notes': notes_map.get(url, ''),
            '_id': ref['_id'],
            'en': ref['en'],
        })

print(f"  Total report rows (occurrences): {len(report_rows)}")

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Broken Links"

NUM_COLS = 11
LAST_COL = get_column_letter(NUM_COLS)

# Column widths
col_widths = [4, 10, 28, 36, 50, 12, 44, 50, 44, 28, 60]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Border helper
thin = Side(style='thin', color='000000')


def thin_border():
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def apply_border(cell):
    cell.border = thin_border()


# ─── Row 1: Title ─────────────────────────────────────────────────────────────
ws.merge_cells(f'A1:{LAST_COL}1')
title_cell = ws['A1']
title_cell.value = (
    f"AnnualUpdates Broken Links Report — {len(broken_urls)} Broken URL(s) Found"
)
title_cell.font = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
title_cell.fill = PatternFill('solid', fgColor='1F3864')
title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws.row_dimensions[1].height = 28
for c in range(1, NUM_COLS + 1):
    apply_border(ws.cell(1, c))

# ─── Row 2: Methodology note ──────────────────────────────────────────────────
ws.merge_cells(f'A2:{LAST_COL}2')
note_cell = ws['A2']
note_cell.value = (
    f"Methodology: Flags only HTTP 404, 410, and DNS resolution failures. "
    f"SSLError sites are retried with verify=False to detect true 404/410. "
    f"Timeouts, 403/401/429/5xx excluded. "
    f"{total} unique URLs scanned across {len(csv_files)} CSV files."
)
note_cell.font = Font(name='Calibri', italic=True, size=9, color='595959')
note_cell.fill = PatternFill('solid', fgColor='F7F7F7')
note_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws.row_dimensions[2].height = 22
for c in range(1, NUM_COLS + 1):
    apply_border(ws.cell(2, c))

# ─── Row 3: Headers ───────────────────────────────────────────────────────────
headers = [
    '#', 'Form ID', 'Form Name', 'Form File', 'Broken URL',
    'HTTP Status', 'CFG URL', 'Suggested Replacement URL',
    'Notes', 'CSV _id', 'CSV en (raw field text)'
]
for col, hdr in enumerate(headers, 1):
    cell = ws.cell(3, col, hdr)
    cell.font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='F4A460')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    apply_border(cell)
ws.row_dimensions[3].height = 20

# ─── Data rows ────────────────────────────────────────────────────────────────
FILL_WHITE = PatternFill('solid', fgColor='FFFFFF')
FILL_BLUE  = PatternFill('solid', fgColor='DCE6F1')
FILL_FORM  = PatternFill('solid', fgColor='FDEBD0')

for row_idx, rd in enumerate(report_rows, 1):
    xl_row = row_idx + 3  # rows 1-3 are header
    fill = FILL_WHITE if row_idx % 2 == 1 else FILL_BLUE

    values = [
        row_idx,
        rd['form_id'],
        rd['form_name'],
        rd['form_file'],
        rd['broken_url'],
        rd['http_status'],
        rd['cfg_url'],
        rd['suggested_replacement'],
        rd['notes'],
        rd['_id'],
        rd['en'],
    ]

    for col, val in enumerate(values, 1):
        cell = ws.cell(xl_row, col, val)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        apply_border(cell)

        # Column-specific styles
        if col == 3:  # Form Name — bold brown on peach
            cell.font = Font(name='Calibri', bold=True, size=10, color='7B3F00')
            cell.fill = FILL_FORM
        elif col == 5:  # Broken URL — red italic
            cell.font = Font(name='Calibri', italic=True, size=9, color='C00000')
        elif col == 6:  # HTTP Status — bold red centered
            cell.font = Font(name='Calibri', bold=True, size=10, color='C00000')
            cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
        elif col == 7:  # CFG URL — blue
            cell.font = Font(name='Calibri', size=9, color='0563C1')
        elif col == 8:  # Suggested Replacement — blue underline
            cell.font = Font(name='Calibri', size=9, color='0563C1', underline='single')
        elif col == 10:  # CSV _id — Courier New 9pt
            cell.font = Font(name='Courier New', size=9)
        elif col == 11:  # CSV en — 8pt
            cell.font = Font(name='Calibri', size=8)
        else:
            cell.font = Font(name='Calibri', size=10)

    ws.row_dimensions[xl_row].height = 80

# Freeze at A4
ws.freeze_panes = 'A4'

# Save
os.makedirs(os.path.dirname(OUT_FILE), exist_ok=True)
wb.save(OUT_FILE)
print(f"\nReport saved to: {OUT_FILE}")

# ─── Summary ─────────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("SUMMARY")
print("=" * 60)
print(f"Total CSVs scanned:        {len(csv_files)}")
print(f"Total unique URLs checked: {total}")
print(f"Broken URLs found:         {len(broken_urls)}")
print(f"Report rows (occurrences): {len(report_rows)}")
if broken_urls:
    print("\nBroken URLs:")
    for u, s in sorted(broken_urls.items()):
        repl = suggested_replacements.get(u, '(no replacement found)')
        print(f"  [{s}] {u}")
        print(f"       -> {repl}")
else:
    print("\nNo broken URLs found!")
print("=" * 60)
