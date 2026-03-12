"""
update_cert_report.py
---------------------
Read all translated ISO certificate Markdown files under
knowledge/Cert/ISO/en-US/ and rebuild (or update) the
ISO_Certificate_Validation_Report.xlsx in report/cert/.

The script is fully idempotent: it reads every *_en-US.md file,
parses key fields with Claude (claude-haiku-4-5), and writes a
fresh report — so running it again after adding new certificates
simply regenerates the whole file with the latest data.

Usage:
    python update_cert_report.py [--en-us-dir <path>] [--report <path>]

    Defaults:
        --en-us-dir  knowledge/Cert/ISO/en-US
        --report     report/cert/ISO_Certificate_Validation_Report.xlsx

Requirements:
    pip install anthropic openpyxl

Environment:
    ANTHROPIC_API_KEY must be set.
"""

import argparse
import json
import os
import sys
import io
from datetime import date
from pathlib import Path

import anthropic
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Fix Windows console encoding
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

# ── Colour palette ─────────────────────────────────────────────────────────────
C_NAVY       = "FF1F3864"
C_WHITE_TEXT = "FFFFFFFF"
C_BLACK_TEXT = "FF000000"
C_LABEL_BG   = "FFD6DCE4"
C_ROW_EVEN   = "FFF2F2F2"
C_ROW_ODD    = "FFFFFFFF"
C_VALID_BG   = "FFE2EFDA";  C_VALID_FG   = "FF375623"
C_EXPIRING_BG= "FFFFF2CC";  C_EXPIRING_FG= "FFFF8C00"
C_EXPIRED_BG = "FFFCE4D6";  C_EXPIRED_FG = "FFC00000"
C_SUBHDR_BG  = "FF2E75B6"

FIELD_ORDER = [
    "Source File", "Company Name", "Country", "Standard",
    "Scope of Certification", "Certificate Number", "Certification Body (CB)",
    "Accreditation Body", "Registered Address", "Issue Date",
    "Certificate Cycle", "Expiry Date", "Days Remaining", "Status",
    "Authorised Signatory", "Primary Verification",
    "Alternative Verification", "How to Verify", "Web Search Evidence",
]

EXTRACT_PROMPT = """\
You are given the full text of an ISO certificate (already OCR-extracted and \
translated to English).  Extract the fields below and return them as a single \
JSON object — no markdown fences, no explanation.

Required fields (use null if not found):
  company_name           – legal company name
  country                – country of the certified organisation
  standard               – e.g. "ISO 9001:2015"
  scope                  – full scope of certification (one paragraph)
  cert_number            – certificate number / reference
  cert_body              – certification body name
  accreditation_body     – accreditation body and number
  registered_address     – registered address of the certified company
  issue_date             – issue / initial certification date (human-readable)
  cert_cycle             – issue number / version / original certification date
  expiry_date            – expiry / valid-to date (human-readable)
  expiry_iso             – expiry date as YYYY-MM-DD (for calculation); null if unknown
  authorised_signatory   – name and title of signatory
  primary_verification   – best URL for online verification
  alt_verification       – secondary URL or method (email, QR, etc.)
  how_to_verify          – short instructions (1-2 sentences)
  web_search_evidence    – brief sentence about CB accreditation / registry evidence

Certificate text:
{text}
"""


# ── Helpers ────────────────────────────────────────────────────────────────────

def fill(rgb):
    return PatternFill(fill_type="solid", fgColor=rgb)


def font(bold=False, color=C_BLACK_TEXT, size=10, name="Calibri"):
    return Font(bold=bold, color=color, size=size, name=name)


def align(h="left", v="top", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def status_style(status: str):
    """Return (fill, font) for VALID / EXPIRING SOON / EXPIRED."""
    s = status.upper()
    if "EXPIRED" in s and "SOON" not in s:
        return fill(C_EXPIRED_BG), font(bold=True, color=C_EXPIRED_FG)
    if "EXPIRING" in s or "SOON" in s:
        return fill(C_EXPIRING_BG), font(bold=True, color=C_EXPIRING_FG)
    return fill(C_VALID_BG), font(bold=True, color=C_VALID_FG)


def days_remaining(expiry_iso: str | None) -> int | None:
    """Return integer days from today to expiry (negative = expired)."""
    if not expiry_iso:
        return None
    try:
        exp = date.fromisoformat(expiry_iso)
        return (exp - date.today()).days
    except ValueError:
        return None


def compute_status(days: int | None) -> str:
    if days is None:
        return "UNKNOWN"
    if days < 0:
        return "EXPIRED"
    if days <= 180:
        return "EXPIRING SOON"
    return "VALID"


def parse_cert(client: anthropic.Anthropic, md_path: Path) -> dict:
    """Use Claude to extract structured fields from a translated cert .md file."""
    text = md_path.read_text(encoding="utf-8")
    resp = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=2048,
        messages=[{
            "role": "user",
            "content": EXTRACT_PROMPT.format(text=text[:12000]),
        }],
    )
    raw = resp.content[0].text.strip()
    # Strip accidental markdown fences
    if raw.startswith("```"):
        raw = raw.split("\n", 1)[1].rsplit("```", 1)[0]
    data = json.loads(raw)
    days = days_remaining(data.get("expiry_iso"))
    status = compute_status(days)
    data["_source_file"] = md_path.name
    data["_days"] = days
    data["_status"] = status
    return data


# ── Excel builders ─────────────────────────────────────────────────────────────

def write_details_sheet(ws, certs: list[dict]):
    ws.title = "Certificate Details"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 64

    # Sheet title row
    c = ws.cell(1, 1, "Detailed Certificate Records \u2014 ISO Certifications")
    c.font = font(bold=True, color=C_WHITE_TEXT, size=13)
    c.fill = fill(C_NAVY)
    c.alignment = align("center", "center")
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 28

    row = 2
    for cert in certs:
        # blank separator (skip before first)
        if row > 2:
            row += 1  # leave blank

        company  = cert.get("company_name") or "Unknown"
        standard = cert.get("standard") or "Unknown"
        status   = cert.get("_status", "UNKNOWN")
        days     = cert.get("_days")

        # Section header
        header_text = f"  {company}  \u2014  {standard}  [{status}]"
        c = ws.cell(row, 1, header_text)
        c.font = font(bold=True, color=C_WHITE_TEXT)
        c.fill = fill(C_NAVY)
        c.alignment = align("left", "center")
        ws.cell(row, 2).fill = fill(C_NAVY)
        ws.row_dimensions[row].height = 22
        row += 1

        days_str = f"{days} days" if days is not None else "N/A"
        if days is not None and days < 0:
            days_str += " (EXPIRED)"

        field_values = [
            ("Source File",            cert.get("_source_file", "")),
            ("Company Name",           cert.get("company_name", "")),
            ("Country",                cert.get("country", "")),
            ("Standard",               cert.get("standard", "")),
            ("Scope of Certification", cert.get("scope", "")),
            ("Certificate Number",     cert.get("cert_number", "")),
            ("Certification Body (CB)",cert.get("cert_body", "")),
            ("Accreditation Body",     cert.get("accreditation_body", "")),
            ("Registered Address",     cert.get("registered_address", "")),
            ("Issue Date",             cert.get("issue_date", "")),
            ("Certificate Cycle",      cert.get("cert_cycle", "")),
            ("Expiry Date",            cert.get("expiry_date", "")),
            ("Days Remaining",         days_str),
            ("Status",                 status),
            ("Authorised Signatory",   cert.get("authorised_signatory", "")),
            ("Primary Verification",   cert.get("primary_verification", "")),
            ("Alternative Verification", cert.get("alt_verification", "")),
            ("How to Verify",          cert.get("how_to_verify", "")),
            ("Web Search Evidence",    cert.get("web_search_evidence", "")),
        ]

        for i, (field, value) in enumerate(field_values, 1):
            is_status_field = field in ("Days Remaining", "Status")
            ca = ws.cell(row, 1, field)
            ca.font = font(bold=True)
            ca.fill = fill(C_LABEL_BG)
            ca.alignment = align()

            cb = ws.cell(row, 2, value)
            cb.alignment = align()
            if is_status_field:
                sfill, sfont = status_style(status)
                cb.fill = sfill
                cb.font = sfont
            else:
                cb.fill = fill(C_ROW_EVEN if i % 2 == 0 else C_ROW_ODD)
                cb.font = font()
            row += 1


def write_summary_sheet(ws, certs: list[dict], source_dir: str):
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 28

    today_str = date.today().strftime("%-d %B %Y") if sys.platform != "win32" else \
                date.today().strftime("%#d %B %Y")
    n = len(certs)

    # Row 1: report title
    c = ws.cell(1, 1, "ISO Certificate Validation Report")
    c.font = font(bold=True, color=C_WHITE_TEXT, size=14)
    c.fill = fill(C_NAVY)
    c.alignment = align("center", "center")
    ws.merge_cells("A1:G1")
    ws.row_dimensions[1].height = 30

    # Row 2: meta
    valid_count     = sum(1 for c in certs if c["_status"] == "VALID")
    expiring_count  = sum(1 for c in certs if c["_status"] == "EXPIRING SOON")
    expired_count   = sum(1 for c in certs if c["_status"] == "EXPIRED")
    unknown_count   = sum(1 for c in certs if c["_status"] == "UNKNOWN")

    ws.cell(2, 1).value = (
        f"Verification Date: {today_str}  |  Certificates: {n}  |  Source: {source_dir}"
    )
    ws.cell(2, 1).font = font(size=10)
    ws.merge_cells("A2:G2")

    # Row 3: blank
    # Row 4: totals header
    totals = [
        (1, f"{n}\nTOTAL",        C_NAVY,        C_WHITE_TEXT, True),
        (2, "VALID",              C_VALID_BG,    C_VALID_FG,   True),
        (3, None,                 None,           None,         False),
        (4, "EXPIRING SOON (<=180 days)", C_EXPIRING_BG, C_EXPIRING_FG, True),
        (5, None,                 None,           None,         False),
        (6, "EXPIRED",            C_EXPIRED_BG,  C_EXPIRED_FG, True),
        (7, None,                 None,           None,         False),
    ]
    for col, val, bg, fg, bold_ in totals:
        if val is None:
            continue
        c = ws.cell(4, col, val)
        c.font = font(bold=bold_, color=fg or C_BLACK_TEXT)
        c.fill = fill(bg) if bg else PatternFill()
        c.alignment = align("center", "center")

    # Row 5: counts
    for col, val, bg, fg, bold_ in [
        (2, valid_count,    C_VALID_BG,    C_VALID_FG,    True),
        (4, expiring_count, C_EXPIRING_BG, C_EXPIRING_FG, True),
        (6, expired_count,  C_EXPIRED_BG,  C_EXPIRED_FG,  True),
    ]:
        c = ws.cell(5, col, val)
        c.font = font(bold=bold_, color=fg)
        c.fill = fill(bg)
        c.alignment = align("center", "center")

    # Row 6: blank
    # Row 7: column headers
    headers = ["Company", "Standard", "Cert Number", "Expiry Date",
               "Days Left", "Status", "Verification Portal"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(7, col, h)
        c.font = font(bold=True, color=C_WHITE_TEXT)
        c.fill = fill(C_NAVY)
        c.alignment = align("center", "center")

    # Rows 8+: data
    for i, cert in enumerate(certs):
        row = 8 + i
        row_bg = C_ROW_ODD if i % 2 == 0 else C_ROW_EVEN
        days = cert.get("_days")
        status = cert["_status"]
        sfill, sfont = status_style(status)

        plain_fill = fill(row_bg)
        plain_font = font(color=C_BLACK_TEXT)

        vals = [
            cert.get("company_name", ""),
            cert.get("standard", ""),
            cert.get("cert_number", ""),
            cert.get("expiry_date", ""),
            days if days is not None else "N/A",
            status,
            cert.get("primary_verification", ""),
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row, col, val)
            if col in (5, 6):  # Days Left + Status get status colour
                c.fill = sfill
                c.font = sfont
            else:
                c.fill = plain_fill
                c.font = plain_font
            c.alignment = align("left", "center")

    # Action notes
    note_row = 8 + len(certs) + 1
    expiring = [c for c in certs if c["_status"] == "EXPIRING SOON"]
    expired  = [c for c in certs if c["_status"] == "EXPIRED"]

    notes = []
    if expiring:
        items = "; ".join(
            f"{c.get('company_name', '?')} ({c.get('cert_number', '?')}, "
            f"expires {c.get('expiry_date', '?')})"
            for c in expiring
        )
        notes.append(
            f"ACTION REQUIRED: {len(expiring)} certificate(s) expiring within 180 days \u2014 "
            f"{items}. Initiate recertification immediately."
        )
    if expired:
        items = "; ".join(
            f"{c.get('company_name', '?')} ({c.get('cert_number', '?')}, "
            f"expired {c.get('expiry_date', '?')})"
            for c in expired
        )
        notes.append(
            f"EXPIRED: {items} \u2014 verify whether renewal has been issued."
        )
    notes.append(
        "NOTE: IAF merged with ILAC into Global Accreditation Cooperation Incorporated "
        "on 01 Jan 2026. Use CB-specific portals or national accreditation bodies for "
        "live verification."
    )

    for note in notes:
        c = ws.cell(note_row, 1, note)
        c.font = font(size=9)
        c.alignment = align("left", "center", wrap=True)
        ws.merge_cells(f"A{note_row}:G{note_row}")
        note_row += 1


def write_guide_sheet(ws):
    ws.title = "Verification Guide"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 30

    c = ws.cell(1, 1, "How to Verify ISO Certificates \u2014 Official Portals & Steps")
    c.font = font(bold=True, color=C_WHITE_TEXT, size=12)
    c.fill = fill(C_NAVY)
    c.alignment = align("left", "center")
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 25

    headers = ["Portal / Method", "Covers", "URL / Access", "Notes"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(2, col, h)
        c.font = font(bold=True, color=C_WHITE_TEXT)
        c.fill = fill(C_SUBHDR_BG)
        c.alignment = align("center", "center")

    rows = [
        ("UKAS CertCheck",
         "All UKAS-accredited CB certs\n(URS, ISOQAR, SGS UK — 300,000+)",
         "certcheck.ukas.com",
         "Free. Search by company, cert number, or CB."),
        ("SGS Certified Client Directory",
         "All SGS-issued certs worldwide",
         "sgs.com/en/certified-clients-and-products/certified-client-directory",
         "Search by cert number or company."),
        ("JQA Registry (Japan)",
         "JQA-issued certs (Japan)",
         "jqa.jp/cgi-bin/06manage/14_touroku/detail_j.html",
         "Search by cert number (e.g. JQA-2075)."),
        ("Bureau Veritas Brazil",
         "BV-issued certs (Brazil / global)",
         "bureauveritas.com.br/certificacao",
         "Contact certificacao@bureauveritas.com or scan QR code."),
        ("SCCM (Netherlands)",
         "SCCM-scheme certs (Netherlands)",
         "sccm.nl",
         "SCCM publishes all certificates in its scheme."),
        ("GCPL / Globus Certifications",
         "GCPL-issued certs",
         "gcert.co",
         "Search by cert number or QR code."),
        ("IAF / RvA / national ABs",
         "Accreditation body registries",
         "iaf.nu  |  rva.nl",
         "Verify CB accreditation status."),
    ]
    for i, row_data in enumerate(rows, 3):
        row_bg = C_ROW_ODD if i % 2 == 1 else C_ROW_EVEN
        for col, val in enumerate(row_data, 1):
            c = ws.cell(i, col, val)
            c.font = font()
            c.fill = fill(row_bg)
            c.alignment = align()


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Rebuild ISO Certificate Validation Report")
    parser.add_argument("--en-us-dir", default=None,
                        help="Path to en-US/ folder with *_en-US.md files")
    parser.add_argument("--report", default=None,
                        help="Output path for the .xlsx report")
    args = parser.parse_args()

    repo_root = Path(__file__).resolve().parents[4]  # …/rag-skill
    en_us_dir = Path(args.en_us_dir) if args.en_us_dir else \
                repo_root / "knowledge" / "Cert" / "ISO" / "en-US"
    report_path = Path(args.report) if args.report else \
                  repo_root / "report" / "cert" / "ISO_Certificate_Validation_Report.xlsx"

    if not en_us_dir.is_dir():
        print(f"ERROR: en-US directory not found: {en_us_dir}", file=sys.stderr)
        sys.exit(1)

    md_files = sorted(en_us_dir.glob("*_en-US.md"))
    if not md_files:
        print(f"No *_en-US.md files found in: {en_us_dir}", file=sys.stderr)
        sys.exit(1)

    print(f"Source : {en_us_dir}")
    print(f"Output : {report_path}")
    print(f"Files  : {len(md_files)}\n")

    client = anthropic.Anthropic()
    certs = []
    for md in md_files:
        print(f"Parsing: {md.name} ...", flush=True)
        try:
            data = parse_cert(client, md)
            certs.append(data)
            print(f"  -> {data.get('company_name', '?')} | {data.get('standard', '?')} | {data['_status']}")
        except Exception as exc:
            print(f"  ERROR: {exc}", file=sys.stderr)

    # Build workbook
    wb = openpyxl.Workbook()
    ws_sum = wb.active
    ws_det = wb.create_sheet()
    ws_guide = wb.create_sheet()

    write_summary_sheet(ws_sum, certs, str(en_us_dir.relative_to(repo_root)))
    write_details_sheet(ws_det, certs)
    write_guide_sheet(ws_guide)

    report_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(report_path)
    print(f"\nSaved: {report_path}")
    print(f"Certs : {len(certs)}  (Valid: {sum(1 for c in certs if c['_status']=='VALID')}  "
          f"Expiring: {sum(1 for c in certs if c['_status']=='EXPIRING SOON')}  "
          f"Expired: {sum(1 for c in certs if c['_status']=='EXPIRED')})")


if __name__ == "__main__":
    main()
