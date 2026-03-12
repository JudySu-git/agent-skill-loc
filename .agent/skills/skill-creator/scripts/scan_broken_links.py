#!/usr/bin/env python3
"""
Broken Link Scanner for Form CSV files.

Scans all formId_*.csv files in a given folder for URLs, checks each URL,
and outputs an Excel report of broken links (HTTP 404, 410, DNS failures only).

Usage:
    python scan_broken_links.py --csv-dir <path> [--forms-xlsx <path>] [--out <path>]
                                [--threads <n>] [--timeout <s>]

Arguments:
    --csv-dir       Folder containing formId_*.csv files to scan (required)
    --forms-xlsx    Path to Forms-data_*.xlsx for CFG URL lookup (optional)
    --out           Output Excel report path (default: broken_links_report.xlsx in csv-dir)
    --threads       Concurrent HTTP threads (default: 8)
    --timeout       HTTP request timeout in seconds (default: 15)

CSV format expected:
    Each CSV must have at minimum an 'en' column with text that may contain URLs.
    Optionally: '_id' column for row identification.

Examples:
    python scan_broken_links.py --csv-dir "knowledge/forms/AnnualUpdates"
    python scan_broken_links.py --csv-dir "knowledge/forms/AnnualUpdates" \\
        --forms-xlsx "knowledge/forms/Forms-data_2026.xlsx" \\
        --out "report/AnnualUpdates_Broken_Links.xlsx"
"""
import sys
# Force UTF-8 output on Windows to handle non-ASCII URLs
if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    except AttributeError:
        pass

import os
import re
import csv
import glob
import argparse
import socket
import requests
import urllib3
import pandas as pd
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

urllib3.disable_warnings()

URL_RE = re.compile(r'https?://[^\s"\'<>)]+')

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
}


# ─── URL helpers ─────────────────────────────────────────────────────────────

def canonicalize_url(url):
    """Strip trailing punctuation, lowercase scheme+host, strip trailing slash."""
    url = url.rstrip('"\'>)')
    m = re.match(r'^(https?://)([^/]+)(.*)', url)
    if m:
        url = (m.group(1) + m.group(2)).lower() + m.group(3)
    return url.rstrip('/')


# ─── HTTP checker ─────────────────────────────────────────────────────────────

def check_url(url, timeout=15):
    """
    Returns (status_code_or_string, final_url).

    Broken = HTTP 404, HTTP 410, or DNS_FAIL.
    SSLError sites are retried with verify=False to get the real HTTP status
    (some servers have self-signed/expired certs but still return meaningful
    HTTP status codes on retry).
    Timeouts, 403/401/429/5xx are NOT flagged as broken (server/bot-blocking).

    Geo-restriction / VPN-gated sites:
      If the full URL is unreachable (DNS_FAIL or ConnErr) but the domain root
      IS reachable, it is likely a genuine 404-equivalent (path removed).
      If even the domain root is unreachable, status is marked UNVERIFIABLE.
    """
    try:
        r = requests.get(url, timeout=timeout, allow_redirects=True, headers=HEADERS)
        return r.status_code, r.url
    except requests.exceptions.SSLError:
        # Retry without SSL verification to get the actual HTTP status
        try:
            r = requests.get(url, timeout=timeout, allow_redirects=True,
                             headers=HEADERS, verify=False)
            return r.status_code, r.url
        except Exception as e2:
            return f"SSL+{type(e2).__name__}", url
    except requests.exceptions.ConnectionError as e:
        if isinstance(e.args[0], urllib3.exceptions.MaxRetryError):
            cause = e.args[0].reason
            if isinstance(cause, urllib3.exceptions.NewConnectionError):
                return _classify_dns_fail(url, timeout), url
        return f"ConnErr:{type(e).__name__}", url
    except requests.exceptions.Timeout:
        return "TIMEOUT", url
    except Exception as e:
        return f"ERR:{type(e).__name__}", url


def _classify_dns_fail(url, timeout):
    """
    On DNS/connection failure for a deep path URL, try the domain root.
    - If domain root also fails → UNVERIFIABLE (possible geo-restriction)
    - If domain root succeeds  → DNS_FAIL (path genuinely gone)
    """
    m = re.match(r'^(https?://[^/]+)', url)
    if not m:
        return "DNS_FAIL"
    domain_root = m.group(1)
    if domain_root == url:
        return "DNS_FAIL"
    try:
        requests.get(domain_root, timeout=timeout, headers=HEADERS)
        return "DNS_FAIL"   # domain reachable → path is gone
    except Exception:
        return "UNVERIFIABLE"   # whole domain unreachable → geo-block/VPN


def is_broken(status):
    return status in (404, 410, "DNS_FAIL")


# ─── CSV scanning ─────────────────────────────────────────────────────────────

def scan_csv_folder(csv_dir):
    """
    Scan all formId_*.csv files in csv_dir.
    Returns url_refs dict: url -> list of {form_id, file, _id, en}
    """
    csv_files = sorted(glob.glob(os.path.join(csv_dir, "formId_*.csv")))
    print(f"  Found {len(csv_files)} CSV files in: {csv_dir}")

    url_refs = defaultdict(list)

    for csv_path in csv_files:
        fname = os.path.basename(csv_path)
        m = re.match(r'formId_(\d+)_', fname)
        form_id = m.group(1) if m else "unknown"

        try:
            with open(csv_path, encoding='utf-8-sig', newline='') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    en_text = row.get('en', '')
                    if not en_text:
                        continue
                    for raw_url in URL_RE.findall(en_text):
                        norm = canonicalize_url(raw_url)
                        if norm:
                            url_refs[norm].append({
                                'form_id': form_id,
                                'file': fname,
                                '_id': row.get('_id', ''),
                                'en': en_text,
                            })
        except Exception as e:
            print(f"  WARNING: Could not read {fname}: {e}")

    return csv_files, url_refs


# ─── CFG URL lookup ───────────────────────────────────────────────────────────

def load_cfg_urls(xlsx_path):
    """Load form_id -> CFG URL mapping from a Forms-data_*.xlsx file."""
    cfg_urls = {}
    if not xlsx_path or not os.path.isfile(xlsx_path):
        return cfg_urls
    try:
        df = pd.read_excel(xlsx_path)
        col_map = {c.strip().lower(): c for c in df.columns}
        form_id_col = next(
            (v for k, v in col_map.items()
             if 'form id' in k or k in ('formid', 'form_id')), None)
        cfg_col = next(
            (v for k, v in col_map.items()
             if 'cfg' in k and 'url' in k), None)
        if form_id_col and cfg_col:
            for _, row in df.iterrows():
                fid = str(row[form_id_col]).strip() if pd.notna(row[form_id_col]) else ''
                curl = str(row[cfg_col]).strip() if pd.notna(row[cfg_col]) else ''
                if fid and fid != 'nan':
                    cfg_urls[fid] = curl
            print(f"  Loaded {len(cfg_urls)} CFG URLs from: {xlsx_path}")
        else:
            print(f"  WARNING: Could not identify Form ID or CFG url columns in {xlsx_path}")
    except Exception as e:
        print(f"  WARNING: Could not read {xlsx_path}: {e}")
    return cfg_urls


# ─── Form name lookup ─────────────────────────────────────────────────────────

def load_form_names(csv_dir):
    """
    Try to load form names from data_structure.md in csv_dir.
    Returns dict: form_id_str -> form_name_str
    """
    form_names = {}
    md_path = os.path.join(csv_dir, "data_structure.md")
    if not os.path.isfile(md_path):
        return form_names
    try:
        with open(md_path, encoding='utf-8') as f:
            for line in f:
                m = re.match(r'\|\s*(\d+)\s*\|\s*([^|]+?)\s*\|', line)
                if m:
                    form_names[m.group(1)] = m.group(2).strip()
        print(f"  Loaded {len(form_names)} form names from data_structure.md")
    except Exception as e:
        print(f"  WARNING: Could not parse data_structure.md: {e}")
    return form_names


# ─── Excel report builder ─────────────────────────────────────────────────────

def build_excel_report(out_path, broken_urls, url_refs, csv_files,
                        total_checked, form_names, cfg_urls,
                        suggested_replacements=None, notes_map=None):
    """Write a formatted Excel report of broken links."""
    if suggested_replacements is None:
        suggested_replacements = {}
    if notes_map is None:
        notes_map = {}

    report_rows = []
    for url, status in sorted(broken_urls.items()):
        seen = set()
        for ref in url_refs.get(url, []):
            key = (ref['form_id'], ref['_id'])
            if key in seen:
                continue
            seen.add(key)
            form_id = ref['form_id']
            report_rows.append({
                'form_id': form_id,
                'form_name': form_names.get(form_id, ''),
                'form_file': ref['file'],
                'broken_url': url,
                'http_status': str(status),
                'cfg_url': cfg_urls.get(form_id, ''),
                'suggested_replacement': suggested_replacements.get(url, ''),
                'notes': notes_map.get(url, ''),
                '_id': ref['_id'],
                'en': ref['en'],
            })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Broken Links"

    NUM_COLS = 11
    LAST_COL = get_column_letter(NUM_COLS)
    thin = Side(style='thin', color='000000')

    def border():
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def apply_border(cell):
        cell.border = border()

    # Column widths: #, Form ID, Form Name, Form File, Broken URL,
    #                Status, CFG URL, Suggested Replacement, Notes, _id, en
    for i, w in enumerate([4, 10, 28, 36, 50, 12, 44, 50, 44, 28, 60], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Row 1: title
    folder_name = os.path.basename(os.path.normpath(
        os.path.dirname(list(url_refs.values())[0][0]['file'])
        if url_refs else ''))
    ws.merge_cells(f'A1:{LAST_COL}1')
    tc = ws['A1']
    tc.value = f"Broken Links Report — {len(broken_urls)} Broken URL(s) Found"
    tc.font = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
    tc.fill = PatternFill('solid', fgColor='1F3864')
    tc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 28
    for c in range(1, NUM_COLS + 1):
        apply_border(ws.cell(1, c))

    # Row 2: methodology note
    ws.merge_cells(f'A2:{LAST_COL}2')
    nc = ws['A2']
    nc.value = (
        f"Methodology: Flags only HTTP 404, 410, and DNS resolution failures. "
        f"SSLError sites retried with verify=False to detect true 404/410. "
        f"UNVERIFIABLE = domain unreachable (possible geo-restriction/VPN). "
        f"Timeouts, 403/401/429/5xx excluded. "
        f"{total_checked} unique URLs scanned across {len(csv_files)} CSV files."
    )
    nc.font = Font(name='Calibri', italic=True, size=9, color='595959')
    nc.fill = PatternFill('solid', fgColor='F7F7F7')
    nc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[2].height = 22
    for c in range(1, NUM_COLS + 1):
        apply_border(ws.cell(2, c))

    # Row 3: headers
    headers = [
        '#', 'Form ID', 'Form Name', 'Form File', 'Broken URL',
        'HTTP Status', 'CFG URL', 'Suggested Replacement URL',
        'Notes', 'CSV _id', 'CSV en (raw field text)'
    ]
    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(3, col, hdr)
        cell.font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='F4A460')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        apply_border(cell)
    ws.row_dimensions[3].height = 20

    # Data rows
    FILL_WHITE = PatternFill('solid', fgColor='FFFFFF')
    FILL_BLUE  = PatternFill('solid', fgColor='DCE6F1')
    FILL_FORM  = PatternFill('solid', fgColor='FDEBD0')

    for row_idx, rd in enumerate(report_rows, 1):
        xl_row = row_idx + 3
        fill = FILL_WHITE if row_idx % 2 == 1 else FILL_BLUE
        values = [
            row_idx, rd['form_id'], rd['form_name'], rd['form_file'],
            rd['broken_url'], rd['http_status'], rd['cfg_url'],
            rd['suggested_replacement'], rd['notes'], rd['_id'], rd['en'],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(xl_row, col, val)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            apply_border(cell)
            if col == 3:
                cell.font = Font(name='Calibri', bold=True, size=10, color='7B3F00')
                cell.fill = FILL_FORM
            elif col == 5:
                cell.font = Font(name='Calibri', italic=True, size=9, color='C00000')
            elif col == 6:
                cell.font = Font(name='Calibri', bold=True, size=10, color='C00000')
                cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
            elif col == 7:
                cell.font = Font(name='Calibri', size=9, color='0563C1')
            elif col == 8:
                cell.font = Font(name='Calibri', size=9, color='0563C1', underline='single')
            elif col == 10:
                cell.font = Font(name='Courier New', size=9)
            elif col == 11:
                cell.font = Font(name='Calibri', size=8)
            else:
                cell.font = Font(name='Calibri', size=10)
        ws.row_dimensions[xl_row].height = 80

    ws.freeze_panes = 'A4'
    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    wb.save(out_path)
    print(f"\nReport saved to: {out_path}")
    return len(report_rows)


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Scan form CSV files for broken links and produce an Excel report."
    )
    parser.add_argument('--csv-dir', required=True,
                        help='Folder containing formId_*.csv files')
    parser.add_argument('--forms-xlsx', default=None,
                        help='Path to Forms-data_*.xlsx for CFG URL lookup')
    parser.add_argument('--out', default=None,
                        help='Output Excel path (default: <csv-dir>/broken_links_report.xlsx)')
    parser.add_argument('--threads', type=int, default=8,
                        help='Concurrent HTTP threads (default: 8)')
    parser.add_argument('--timeout', type=int, default=15,
                        help='HTTP request timeout in seconds (default: 15)')
    args = parser.parse_args()

    csv_dir = os.path.abspath(args.csv_dir)
    out_path = args.out or os.path.join(csv_dir, "broken_links_report.xlsx")

    print("=" * 60)
    print("STEP 1: Scanning CSVs for URLs …")
    csv_files, url_refs = scan_csv_folder(csv_dir)
    unique_urls = list(url_refs.keys())
    print(f"  Unique normalized URLs found: {len(unique_urls)}")

    print(f"\nSTEP 2: Checking URLs ({args.threads} threads, timeout={args.timeout}s) …")
    results = {}
    done = 0
    total = len(unique_urls)

    with ThreadPoolExecutor(max_workers=args.threads) as ex:
        futures = {ex.submit(check_url, u, args.timeout): u for u in unique_urls}
        for fut in as_completed(futures):
            orig_url = futures[fut]
            status, final_url = fut.result()
            results[orig_url] = (status, final_url)
            done += 1
            print(f"  [{done:3}/{total}] {status}  {orig_url}")

    print("\nSTEP 3: Filtering broken URLs (404, 410, DNS_FAIL) …")
    broken_urls = {url: status for url, (status, _) in results.items() if is_broken(status)}
    print(f"  Total URLs checked:          {total}")
    print(f"  Broken (404/410/DNS_FAIL):   {len(broken_urls)}")
    for u, s in sorted(broken_urls.items()):
        print(f"    [{s}] {u}")

    if not broken_urls:
        print("\nNo broken URLs found. No report generated.")
        return

    print("\nSTEP 4: Enriching data …")
    form_names = load_form_names(csv_dir)
    cfg_urls = load_cfg_urls(args.forms_xlsx)

    print("\nSTEP 5: Building Excel report …")
    # suggested_replacements and notes_map are empty by default.
    # To add replacements, either:
    #   (a) edit this script and populate the dicts below, or
    #   (b) import and call build_excel_report() directly from another script.
    num_rows = build_excel_report(
        out_path=out_path,
        broken_urls=broken_urls,
        url_refs=url_refs,
        csv_files=csv_files,
        total_checked=total,
        form_names=form_names,
        cfg_urls=cfg_urls,
    )

    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"CSVs scanned:           {len(csv_files)}")
    print(f"Unique URLs checked:    {total}")
    print(f"Broken URLs found:      {len(broken_urls)}")
    print(f"Report rows:            {num_rows}")
    print(f"Output:                 {out_path}")
    print("=" * 60)


if __name__ == '__main__':
    main()
